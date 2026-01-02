"""
Excel export service for generating comprehensive financial models
Creates a 5-sheet Excel workbook with deal analysis
"""
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from app.services.deal_service import DealService
from app.services.deal_memo_service import DealMemoService


class ExcelExportService:
    """Service for generating Excel financial models"""

    # Color scheme
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    TITLE_FONT = Font(bold=True, size=14)
    BOLD_FONT = Font(bold=True)

    BORDER_THIN = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @staticmethod
    def generate_excel(deal_id: int):
        """
        Generate Excel workbook for a deal

        Args:
            deal_id: ID of the deal to export

        Returns:
            BytesIO object containing Excel file
        """
        deal_model = DealService.get_deal_model(deal_id)
        if not deal_model:
            return None

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        ExcelExportService._create_executive_summary(wb, deal_model)
        ExcelExportService._create_assumptions_sheet(wb, deal_model)
        ExcelExportService._create_cash_flow_sheet(wb, deal_model)
        ExcelExportService._create_market_data_sheet(wb, deal_model)
        ExcelExportService._create_returns_analysis(wb, deal_model)

        # Add risk assessment sheets if available
        try:
            ExcelExportService._create_risk_assessment_sheet(wb, deal_model)
            ExcelExportService._create_deal_memo_sheet(wb, deal_model)
            ExcelExportService._create_sensitivity_analysis_sheet(wb, deal_model)
        except Exception as e:
            # Risk assessment may not be available for all deals
            print(f"Note: Risk assessment sheets not added: {str(e)}")

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    @staticmethod
    def _create_executive_summary(wb, deal):
        """Create Executive Summary sheet"""
        ws = wb.create_sheet("Executive Summary")

        # Title
        ws['A1'] = deal.deal_name
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Status: {deal.status.upper()}"
        ws['A2'].font = ExcelExportService.BOLD_FONT

        row = 4

        # Deal Information
        ws[f'A{row}'] = "DEAL INFORMATION"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        info_items = [
            ("Location:", deal.location or "N/A"),
            ("Property Address:", deal.property_address or "N/A"),
            ("Property Type:", deal.property_type or "N/A"),
            ("Bedrooms:", deal.bedrooms or "N/A"),
            ("Bathrooms:", deal.bathrooms or "N/A"),
            ("Square Footage:", f"{deal.square_footage:,}" if deal.square_footage else "N/A"),
        ]

        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = ExcelExportService.BOLD_FONT
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # Financial Summary
        ws[f'A{row}'] = "FINANCIAL SUMMARY"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        financial_items = [
            ("Purchase Price:", deal.purchase_price, "currency"),
            ("Monthly Rent:", deal.monthly_rent, "currency"),
            ("Monthly Cash Flow:", deal.monthly_cash_flow, "currency"),
            ("Cash-on-Cash Return:", deal.cash_on_cash_return, "percent"),
            ("Cap Rate:", deal.cap_rate, "percent"),
            ("ROI:", deal.roi, "percent"),
        ]

        for label, value, format_type in financial_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = ExcelExportService.BOLD_FONT

            if value is not None:
                ws[f'B{row}'] = value
                if format_type == "currency":
                    ws[f'B{row}'].number_format = '"$"#,##0.00'
                elif format_type == "percent":
                    ws[f'B{row}'].number_format = '0.00"%"'

                # Color code cash flow
                if label == "Monthly Cash Flow:":
                    if value > 0:
                        ws[f'B{row}'].fill = ExcelExportService.POSITIVE_FILL
                    elif value < 0:
                        ws[f'B{row}'].fill = ExcelExportService.NEGATIVE_FILL
            else:
                ws[f'B{row}'] = "N/A"

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    @staticmethod
    def _create_assumptions_sheet(wb, deal):
        """Create Assumptions sheet with all input parameters"""
        ws = wb.create_sheet("Assumptions")

        # Title
        ws['A1'] = "Financial Assumptions"
        ws['A1'].font = ExcelExportService.TITLE_FONT
        ws.merge_cells('A1:C1')

        row = 3

        # Purchase Details
        ws[f'A{row}'] = "PURCHASE DETAILS"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        purchase_items = [
            ("Purchase Price", deal.purchase_price, '"$"#,##0.00'),
            ("Down Payment %", deal.down_payment_percent, '0.00"%"'),
            ("Loan Interest Rate %", deal.loan_interest_rate, '0.00"%"'),
            ("Loan Term (Years)", deal.loan_term_years, '0'),
            ("Closing Costs", deal.closing_costs, '"$"#,##0.00'),
        ]

        for label, value, number_format in purchase_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value if value is not None else 0
            ws[f'B{row}'].number_format = number_format
            row += 1

        row += 1

        # Income
        ws[f'A{row}'] = "INCOME"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        income_items = [
            ("Monthly Rent", deal.monthly_rent, '"$"#,##0.00'),
            ("Other Monthly Income", deal.other_monthly_income, '"$"#,##0.00'),
            ("Vacancy Rate %", deal.vacancy_rate, '0.00"%"'),
            ("Annual Rent Increase %", deal.annual_rent_increase, '0.00"%"'),
        ]

        for label, value, number_format in income_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value if value is not None else 0
            ws[f'B{row}'].number_format = number_format
            row += 1

        row += 1

        # Expenses
        ws[f'A{row}'] = "EXPENSES"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        expense_items = [
            ("Property Tax (Annual)", deal.property_tax_annual, '"$"#,##0.00'),
            ("Insurance (Annual)", deal.insurance_annual, '"$"#,##0.00'),
            ("HOA (Monthly)", deal.hoa_monthly, '"$"#,##0.00'),
            ("Maintenance %", deal.maintenance_percent, '0.00"%"'),
            ("Property Management %", deal.property_management_percent, '0.00"%"'),
            ("Utilities (Monthly)", deal.utilities_monthly, '"$"#,##0.00'),
            ("Other Expenses (Monthly)", deal.other_expenses_monthly, '"$"#,##0.00'),
        ]

        for label, value, number_format in expense_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value if value is not None else 0
            ws[f'B{row}'].number_format = number_format
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    @staticmethod
    def _create_cash_flow_sheet(wb, deal):
        """Create Cash Flow Analysis sheet with chart"""
        ws = wb.create_sheet("Cash Flow")

        # Title
        ws['A1'] = "30-Year Cash Flow Projection"
        ws['A1'].font = ExcelExportService.TITLE_FONT
        ws.merge_cells('A1:E1')

        # Headers
        headers = ["Year", "Income", "Expenses", "Debt Service", "Cash Flow"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = ExcelExportService.HEADER_FONT
            cell.fill = ExcelExportService.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        # Calculate cash flow for 30 years
        monthly_rent = deal.monthly_rent or 0
        other_income = deal.other_monthly_income or 0
        monthly_payment = deal.monthly_payment or 0
        total_expenses = deal.total_monthly_expenses or 0
        annual_rent_increase = (deal.annual_rent_increase or 0) / 100

        for year in range(1, 31):
            row = year + 3

            # Income with annual increase
            annual_income = (monthly_rent + other_income) * 12 * (1 + annual_rent_increase) ** (year - 1)

            # Expenses (assuming 2% annual increase)
            annual_expenses = total_expenses * 12 * (1.02) ** (year - 1)

            # Debt service (constant)
            annual_debt_service = monthly_payment * 12

            # Cash flow
            cash_flow = annual_income - annual_expenses - annual_debt_service

            ws[f'A{row}'] = year
            ws[f'B{row}'] = annual_income
            ws[f'B{row}'].number_format = '"$"#,##0'
            ws[f'C{row}'] = annual_expenses
            ws[f'C{row}'].number_format = '"$"#,##0'
            ws[f'D{row}'] = annual_debt_service
            ws[f'D{row}'].number_format = '"$"#,##0'
            ws[f'E{row}'] = cash_flow
            ws[f'E{row}'].number_format = '"$"#,##0'

            # Color code cash flow
            if cash_flow > 0:
                ws[f'E{row}'].fill = ExcelExportService.POSITIVE_FILL
            elif cash_flow < 0:
                ws[f'E{row}'].fill = ExcelExportService.NEGATIVE_FILL

        # Create chart
        chart = BarChart()
        chart.title = "Annual Cash Flow"
        chart.y_axis.title = "Amount ($)"
        chart.x_axis.title = "Year"

        data = Reference(ws, min_col=5, min_row=3, max_row=33)
        cats = Reference(ws, min_col=1, min_row=4, max_row=33)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "G3")

        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15

    @staticmethod
    def _create_market_data_sheet(wb, deal):
        """Create Market Data sheet with RentCast and FRED data"""
        ws = wb.create_sheet("Market Data")

        # Title
        ws['A1'] = "Market Data & Comparables"
        ws['A1'].font = ExcelExportService.TITLE_FONT
        ws.merge_cells('A1:D1')

        row = 3

        # RentCast Data
        ws[f'A{row}'] = "RENTCAST DATA"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        if deal.rentcast_data:
            try:
                rentcast = json.loads(deal.rentcast_data)
                for key, value in rentcast.items():
                    ws[f'A{row}'] = str(key).replace('_', ' ').title()
                    ws[f'B{row}'] = str(value)
                    row += 1
            except:
                ws[f'A{row}'] = "Data available in deal record"
                row += 1
        else:
            ws[f'A{row}'] = "No RentCast data available"
            row += 1

        row += 1

        # FRED Data
        ws[f'A{row}'] = "FRED ECONOMIC DATA"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        if deal.fred_data:
            try:
                fred = json.loads(deal.fred_data)
                for key, value in fred.items():
                    ws[f'A{row}'] = str(key).replace('_', ' ').title()
                    ws[f'B{row}'] = str(value)
                    row += 1
            except:
                ws[f'A{row}'] = "Data available in deal record"
                row += 1
        else:
            ws[f'A{row}'] = "No FRED data available"
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

    @staticmethod
    def _create_returns_analysis(wb, deal):
        """Create Returns Analysis sheet"""
        ws = wb.create_sheet("Returns Analysis")

        # Title
        ws['A1'] = "Investment Returns Analysis"
        ws['A1'].font = ExcelExportService.TITLE_FONT
        ws.merge_cells('A1:C1')

        row = 3

        # Key Metrics
        ws[f'A{row}'] = "KEY PERFORMANCE METRICS"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        metrics = [
            ("Cash-on-Cash Return", deal.cash_on_cash_return, '0.00"%"'),
            ("Cap Rate", deal.cap_rate, '0.00"%"'),
            ("Return on Investment (ROI)", deal.roi, '0.00"%"'),
            ("Net Present Value (NPV)", deal.npv, '"$"#,##0.00'),
            ("Internal Rate of Return (IRR)", deal.irr, '0.00"%"'),
        ]

        for label, value, number_format in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = ExcelExportService.BOLD_FONT

            if value is not None:
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = number_format

                # Color code based on value
                if value > 0:
                    ws[f'B{row}'].fill = ExcelExportService.POSITIVE_FILL
                elif value < 0:
                    ws[f'B{row}'].fill = ExcelExportService.NEGATIVE_FILL
            else:
                ws[f'B{row}'] = "N/A"

            row += 1

        row += 2

        # Monthly Breakdown
        ws[f'A{row}'] = "MONTHLY BREAKDOWN"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        monthly_items = [
            ("Total Monthly Income", deal.total_monthly_income, '"$"#,##0.00'),
            ("Total Monthly Expenses", deal.total_monthly_expenses, '"$"#,##0.00'),
            ("Monthly Debt Service", deal.monthly_payment, '"$"#,##0.00'),
            ("Monthly Cash Flow", deal.monthly_cash_flow, '"$"#,##0.00'),
        ]

        for label, value, number_format in monthly_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = ExcelExportService.BOLD_FONT

            if value is not None:
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = number_format
            else:
                ws[f'B{row}'] = 0
                ws[f'B{row}'].number_format = number_format

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    @staticmethod
    def _create_risk_assessment_sheet(wb, deal):
        """Create Risk Assessment sheet with comprehensive analysis"""
        ws = wb.create_sheet("Risk Assessment")

        # Get risk assessment data
        assessment = DealService.get_risk_assessment(deal.id)
        if not assessment:
            ws['A1'] = "No risk assessment available for this deal"
            ws['A1'].font = ExcelExportService.TITLE_FONT
            return

        # Title
        ws['A1'] = "Risk Assessment Analysis"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # Deal name
        ws['A2'] = deal.deal_name
        ws['A2'].font = Font(bold=True, size=12, color="4472C4")
        ws.merge_cells('A2:D2')

        row = 4

        # Section 1: Rent Tier Classification
        ws[f'A{row}'] = "RENT TIER CLASSIFICATION"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        tier_data = [
            ("Rent Tier", assessment.get('rent_tier_label', 'N/A')),
            ("National Decile", assessment.get('rent_decile_national', 'N/A')),
            ("Regional Decile", assessment.get('rent_decile_regional', 'N/A')),
            ("Predicted Rent (Monthly)", f"${assessment.get('predicted_fundamental_rent', 0):.2f}"),
            ("Percentile", f"{assessment.get('rent_percentile', 0):.1f}%"),
        ]

        for label, value in tier_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = ExcelExportService.BOLD_FONT
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # Section 2: Yield Analysis
        ws[f'A{row}'] = "YIELD ANALYSIS"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        yield_data = [
            ("Gross Yield", f"{assessment.get('gross_yield', 0):.2f}%"),
            ("Maintenance Cost", f"{assessment.get('maintenance_cost_pct', 0):.2f}%"),
            ("Property Tax", f"{assessment.get('property_tax_pct', 0):.2f}%"),
            ("Turnover Cost", f"{assessment.get('turnover_cost_pct', 0):.2f}%"),
            ("Default Cost", f"{assessment.get('default_cost_pct', 0):.2f}%"),
            ("Management Cost", f"{assessment.get('management_cost_pct', 0):.2f}%"),
            ("Net Yield", f"{assessment.get('net_yield', 0):.2f}%"),
            ("vs Benchmark", assessment.get('vs_benchmark_yield', 'N/A')),
        ]

        for label, value in yield_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = ExcelExportService.BOLD_FONT if 'Net Yield' in label else None
            ws[f'B{row}'] = value
            if 'Net Yield' in label:
                ws[f'B{row}'].fill = ExcelExportService.POSITIVE_FILL
            row += 1

        row += 1

        # Section 3: Total Returns
        ws[f'A{row}'] = "TOTAL RETURNS"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        return_data = [
            ("Total Return (Unlevered)", f"{assessment.get('total_return_unlevered', 0):.2f}%"),
            ("Total Return (Levered)", f"{assessment.get('total_return_levered', 0):.2f}%"),
            ("Capital Gain (Annual)", f"{assessment.get('capital_gain_yield_annual', 0):.2f}%"),
            ("vs Benchmark", assessment.get('vs_benchmark_return', 'N/A')),
        ]

        for label, value in return_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = ExcelExportService.BOLD_FONT if 'Levered' in label else None
            ws[f'B{row}'] = value
            if 'Levered' in label:
                ws[f'B{row}'].fill = ExcelExportService.POSITIVE_FILL
            row += 1

        row += 1

        # Section 4: Risk Scores
        ws[f'A{row}'] = "RISK ANALYSIS"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        risk_data = [
            ("Systematic Risk Score", f"{assessment.get('systematic_risk_score', 0):.1f}/100"),
            ("Regulatory Risk Score", f"{assessment.get('regulatory_risk_score', 0):.1f}/100"),
            ("Idiosyncratic Risk Score", f"{assessment.get('idiosyncratic_risk_score', 0):.1f}/100"),
            ("Composite Risk Score", f"{assessment.get('composite_risk_score', 0):.1f}/100"),
            ("Risk Level", assessment.get('composite_risk_level', 'N/A')),
        ]

        for label, value in risk_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = ExcelExportService.BOLD_FONT if 'Composite' in label else None
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # Section 5: Arbitrage Opportunity
        ws[f'A{row}'] = "ARBITRAGE OPPORTUNITY"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        arbitrage_data = [
            ("Arbitrage Score", f"{assessment.get('arbitrage_opportunity_score', 0):.1f}/100"),
            ("Opportunity Level", assessment.get('arbitrage_opportunity_level', 'N/A')),
            ("Recommended Investor", assessment.get('recommended_investor_type', 'N/A')),
        ]

        for label, value in arbitrage_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = ExcelExportService.BOLD_FONT if 'Score' in label else None
            ws[f'B{row}'] = value
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    @staticmethod
    def _create_deal_memo_sheet(wb, deal):
        """Create Deal Memo sheet with investment recommendation"""
        ws = wb.create_sheet("Deal Memo")

        # Get deal memo data
        try:
            memo = DealMemoService.generate_memo(deal.id)
        except Exception as e:
            ws['A1'] = f"Error generating deal memo: {str(e)}"
            return

        # Title
        ws['A1'] = "Investment Memo"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        row = 3

        # Executive Summary
        ws[f'A{row}'] = "EXECUTIVE SUMMARY"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        exec_summary = memo.get('executive_summary', {})
        exec_data = [
            ("Property", exec_summary.get('property', 'N/A')),
            ("Address", exec_summary.get('address', 'N/A')),
            ("Purchase Price", f"${exec_summary.get('purchase_price', 0):,.0f}"),
            ("Rent Tier", exec_summary.get('rent_tier', 'N/A')),
            ("Tier Category", exec_summary.get('tier_category', 'N/A')),
            ("Expected Return (Levered)", f"{exec_summary.get('calculated_return_levered', 0):.2f}%"),
            ("Risk Level", exec_summary.get('risk_level', 'N/A')),
            ("Overall Rating", exec_summary.get('overall_rating', 'N/A')),
            ("Target Investor", exec_summary.get('target_investor', 'N/A')),
        ]

        for label, value in exec_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = ExcelExportService.BOLD_FONT
            ws[f'B{row}'] = value
            if 'Rating' in label:
                ws[f'B{row}'].fill = ExcelExportService.POSITIVE_FILL
            row += 1

        row += 2

        # Investment Recommendation
        ws[f'A{row}'] = "INVESTMENT RECOMMENDATION"
        ws[f'A{row}'].font = ExcelExportService.HEADER_FONT
        ws[f'A{row}'].fill = ExcelExportService.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        recommendation = memo.get('investment_recommendation', {})
        ws[f'A{row}'] = "Rating"
        ws[f'A{row}'].font = ExcelExportService.BOLD_FONT
        ws[f'B{row}'] = recommendation.get('overall_rating', 'N/A')
        ws[f'B{row}'].fill = ExcelExportService.POSITIVE_FILL
        row += 1

        ws[f'A{row}'] = "Score"
        ws[f'A{row}'].font = ExcelExportService.BOLD_FONT
        ws[f'B{row}'] = f"{recommendation.get('rating_score', 0)}/100"
        row += 2

        # Key Strengths
        key_strengths = recommendation.get('key_strengths', [])
        if key_strengths:
            ws[f'A{row}'] = "Key Strengths:"
            ws[f'A{row}'].font = Font(bold=True, color="006100")
            row += 1
            for strength in key_strengths:
                ws[f'A{row}'] = f"✓ {strength}"
                ws.merge_cells(f'A{row}:D{row}')
                row += 1

        row += 1

        # Key Concerns
        key_concerns = recommendation.get('key_concerns', [])
        if key_concerns:
            ws[f'A{row}'] = "Key Concerns:"
            ws[f'A{row}'].font = Font(bold=True, color="9C0006")
            row += 1
            for concern in key_concerns:
                ws[f'A{row}'] = f"⚠ {concern}"
                ws.merge_cells(f'A{row}:D{row}')
                row += 1

        row += 2

        # Summary
        ws[f'A{row}'] = "Summary:"
        ws[f'A{row}'].font = ExcelExportService.BOLD_FONT
        row += 1
        ws[f'A{row}'] = recommendation.get('summary', '')
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    @staticmethod
    def _create_sensitivity_analysis_sheet(wb, deal):
        """Create Sensitivity Analysis sheet"""
        ws = wb.create_sheet("Sensitivity Analysis")

        # Get deal memo for sensitivity data
        try:
            memo = DealMemoService.generate_memo(deal.id)
            sensitivity = memo.get('sensitivity_analysis', {})
        except Exception as e:
            ws['A1'] = f"Error generating sensitivity analysis: {str(e)}"
            return

        # Title
        ws['A1'] = "Sensitivity Analysis"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')

        row = 3

        # Headers
        headers = ['Scenario', 'Rent Assumption', 'Appreciation', 'Net Yield', 'Unlevered Return', 'Levered Return']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = ExcelExportService.HEADER_FONT
            ws.cell(row=row, column=col).fill = ExcelExportService.HEADER_FILL

        row += 1

        # Scenario data
        scenarios = sensitivity.get('scenarios', {})
        for scenario_key, scenario_data in scenarios.items():
            ws.cell(row=row, column=1, value=scenario_data.get('name', scenario_key))
            ws.cell(row=row, column=2, value=f"${scenario_data.get('rent_assumption', 0):,.0f}")
            ws.cell(row=row, column=3, value=f"{scenario_data.get('appreciation_assumption', 0):.2f}%")
            ws.cell(row=row, column=4, value=f"{scenario_data.get('net_yield', 0):.2f}%")
            ws.cell(row=row, column=5, value=f"{scenario_data.get('total_return_unlevered', 0):.2f}%")
            ws.cell(row=row, column=6, value=f"{scenario_data.get('total_return_levered', 0):.2f}%")

            # Highlight base case
            if 'base' in scenario_data.get('name', '').lower():
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = ExcelExportService.SUBHEADER_FILL

            row += 1

        row += 2

        # Interpretation
        ws[f'A{row}'] = "Interpretation:"
        ws[f'A{row}'].font = ExcelExportService.BOLD_FONT
        row += 1
        ws[f'A{row}'] = sensitivity.get('interpretation', '')
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)

        # Column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
